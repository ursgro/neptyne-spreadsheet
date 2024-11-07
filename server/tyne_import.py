import re
from datetime import datetime, time
from io import BytesIO
from typing import cast

import gspread
import gspread_asyncio
from google.oauth2.credentials import Credentials
from gspread_formatting import (
    Borders as GSBorder,
)
from gspread_formatting import (
    CellFormat,
    TextFormat,
)
from gspread_formatting import (
    Color as GSColor,
)
from openpyxl import load_workbook
from openpyxl.cell.cell import (
    TYPE_ERROR,
    TYPE_FORMULA,
    TYPE_NUMERIC,
)
from openpyxl.styles import Border as ExcelBorder
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.styles.colors import Color as ExcelColor
from openpyxl.styles.fonts import Font as ExcelFont
from openpyxl.utils.escape import unescape
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session
from tornado.web import HTTPError

from neptyne_kernel import expression_compiler
from neptyne_kernel.cell_address import Address
from neptyne_kernel.expression_compiler import (
    DEFAULT_N_COLS,
    DEFAULT_N_ROWS,
)
from neptyne_kernel.mime_types import JSON_MIME_KEY
from neptyne_kernel.neptyne_protocol import (
    BorderType,
    CellAttribute,
    LineWrap,
    NumberFormat,
    SheetAttribute,
    TextStyle,
    VerticalAlign,
)
from neptyne_kernel.spreadsheet_datetime import (
    SpreadsheetDate,
    SpreadsheetDateTime,
    SpreadsheetTime,
)
from neptyne_kernel.tyne_model.cell import NotebookCell, SheetCell
from neptyne_kernel.tyne_model.jupyter_notebook import (
    CellType,
    Output,
    OutputType,
    jupyter_notebook_from_dict,
)
from neptyne_kernel.tyne_model.sheet import Sheet as TyneSheet
from server.models import User
from server.openpyxl_color_loader import get_theme_colors, theme_and_tint_to_rgb
from server.tyne_content import TyneContent, TyneModelWithContent

TIME_RE = re.compile(r"\d{1,2}:\d{1,2}")
XLFN_RE = re.compile(r"_xlfn(\.\w+)+\(")

ALLOWED_FONTS = (
    "Sans Serif",
    "Noto Serif",
    "Roboto Mono",
    "Sofia Sans",
    "Comic Neue",
    "Biryani",
    "Bodoni Moda",
    "Cedarville",
    "EB Garamond",
    "Libre Franklin",
    "Montserrat",
    "Open Sans",
    "Secular One",
)

BORDER_FIELDS = {
    "top": "BORDER_TOP",
    "bottom": "BORDER_BOTTOM",
    "left": "BORDER_LEFT",
    "right": "BORDER_RIGHT",
}

OPENPYXL_COL_PIXELS_PER_UNIT = 7.5


def get_border_data(border: GSBorder | ExcelBorder) -> str | None:
    if border:
        check_border = (
            (lambda attr: getattr(border, attr))
            if isinstance(border, GSBorder)
            else lambda attr: (border_attr := getattr(border, attr))
            and (hasattr(border_attr, "border_style") and border_attr.border_style)
        )
        return " ".join(
            getattr(BorderType, border_type_attr).value
            for attr, border_type_attr in BORDER_FIELDS.items()
            if check_border(attr)
        )


GS_FONT_STYLE_ATTRS = {
    "bold": TextStyle.BOLD.value,
    "italic": TextStyle.ITALIC.value,
    "underline": TextStyle.UNDERLINE.value,
}

EXCEL_FONT_STYLE_ATTRS = {
    "b": TextStyle.BOLD.value,
    "i": TextStyle.ITALIC.value,
    "u": TextStyle.UNDERLINE.value,
}


def get_font_style(text_fmt: ExcelFont | TextFormat, is_excel: bool) -> str:
    attrs = EXCEL_FONT_STYLE_ATTRS if is_excel else GS_FONT_STYLE_ATTRS
    return " ".join(value for attr, value in attrs.items() if getattr(text_fmt, attr))


def calc_grid_size(column_count: int, row_count: int) -> tuple[int, int]:
    return (
        max(DEFAULT_N_COLS, column_count),
        max(DEFAULT_N_ROWS, row_count),
    )


def get_openpyxl_color(color: ExcelColor, theme_colors: list[str]) -> str | None:
    result = ""
    if color.type == "rgb":
        result = color.rgb
    elif color.type == "theme":
        try:
            result = theme_and_tint_to_rgb(theme_colors, color.theme, color.tint)
        except IndexError:
            return None
    elif color.type == "indexed":
        try:
            result = COLOR_INDEX[color.indexed]
        except IndexError:
            return None

    if (rgb := result[2:] if len(result) == 8 else result) != "000000":
        return "#" + rgb


class TyneImportMixin:
    async def import_tyne_json(
        self, session: Session, tyne_json: dict, file_name: str, user: User
    ) -> TyneModelWithContent:
        content = TyneContent.empty()

        for sheet in tyne_json["sheets"]:
            sheet_id = sheet.get("sheet_id", 0)
            name = sheet.get("name", f"Sheet{sheet_id}")
            tyne_sheet = TyneSheet(sheet_id=sheet_id, name=name)
            tyne_sheet.attributes = sheet.get("attributes", {})
            tyne_sheet.grid_size = (
                sheet.get("n_cols", DEFAULT_N_COLS),
                sheet.get("n_rows", DEFAULT_N_ROWS),
            )
            content.sheets.sheets[sheet_id] = tyne_sheet
            for coord, cell in sheet["contents"].items():
                tyne_sheet.cells[Address.from_a1_or_str(coord)] = SheetCell.from_dict(
                    cell
                )

        content.notebook_cells = []
        requirements = []
        for notebook in tyne_json["notebooks"]:
            if notebook["requirements"]:
                requirements.append(notebook["requirements"])
            for cell in notebook["contents"].values():
                content.notebook_cells.append(NotebookCell.from_dict(cell))

        new_tyne = await self.new_tyne(  # type: ignore
            session,
            user,
            name=tyne_json["name"],
            content=content,
            requirements="\n".join(requirements),
        )
        new_tyne.tyne_model.properties = tyne_json.get("properties", {})

        return new_tyne

    async def import_notebook_ipynb(
        self, session: Session, notebook_json: dict, file_name: str, user: User
    ) -> TyneModelWithContent:
        """Import a jupyter notebook from its json representation as a dict"""
        tyne_content = TyneContent.empty()

        notebook = jupyter_notebook_from_dict(notebook_json)

        def as_str(list_or_str: str | list[str]) -> str:
            if isinstance(list_or_str, list):
                return "".join(list_or_str)
            return list_or_str

        for idx, cell in enumerate(notebook.cells):
            cell_id = f"0{idx}"
            raw_code = as_str(cell.source)
            outputs = cell.outputs
            compiled_code = ""
            if cell.cell_type == CellType.CODE:
                try:
                    compiled_code = expression_compiler.compile_expression(
                        raw_code
                    ).compiled_code
                except ValueError as err:
                    if outputs is None:
                        outputs = []
                    outputs.append(
                        Output(
                            data=None,
                            execution_count=1,
                            metadata=None,
                            output_type=OutputType.ERROR,
                            name=None,
                            text=None,
                            ename=str(type(err)),
                            evalue=str(err),
                            traceback=None,
                        )
                    )
            imported_cell = NotebookCell(
                cell_id=cell_id,
                outputs=outputs,
                raw_code=raw_code,
                compiled_code=compiled_code,
                execution_count=cell.execution_count or 0,
            )
            tyne_content.notebook_cells.append(imported_cell)

        tyne_content.clear_outputs()
        return await self.new_tyne(session, user, name=file_name, content=tyne_content)  # type: ignore

    async def import_google_sheet(
        self, session: Session, url: str, user: User, credentials: Credentials
    ) -> TyneModelWithContent:
        try:
            gc = await gspread_asyncio.AsyncioGspreadClientManager(
                lambda: credentials
            ).authorize()
            wb = await gc.open_by_url(url)
        except gspread.exceptions.NoValidUrlKeyFound:
            raise HTTPError(400, reason="Invalid url")
        except gspread.exceptions.APIError as e:
            raise HTTPError(
                400, reason=f"Google Sheets API error. {e.args[0].get('message')}"
            )

        new_tyne_content = TyneContent.empty()

        wb_data = await wb.fetch_sheet_metadata(
            {
                "fields": "sheets",
                "includeGridData": True,
            }
        )

        for ind, sh in enumerate(wb_data["sheets"]):
            props = sh["properties"]
            data = sh["data"][0]

            if not ind:
                new_tyne_content.sheets.rename_sheet(ind, props["title"])
                sheet = new_tyne_content.sheets.sheets[ind]
            else:
                name, sheet = new_tyne_content.sheets.new_sheet(props["title"])

            grid_props = props["gridProperties"]

            sheet.grid_size = calc_grid_size(
                grid_props["columnCount"], grid_props["rowCount"]
            )

            column_widths = (column["pixelSize"] for column in data["columnMetadata"])
            row_heights = (row["pixelSize"] for row in data["rowMetadata"])

            sheet.attributes[SheetAttribute.ROWS_SIZES.value] = {
                str(i): round(height) for i, height in enumerate(row_heights)
            }
            sheet.attributes[SheetAttribute.COLS_SIZES.value] = {
                str(i): round(width) for i, width in enumerate(column_widths)
            }

            for i, row in enumerate(data.get("rowData", [])):
                if row and (row_values := row.get("values")):
                    for j, cell_value in enumerate(row_values):
                        if not cell_value:
                            continue
                        addr = Address(j, i, sheet.id)
                        attrs = {}
                        user_entered_fmt = cell_value.get("userEnteredFormat")
                        cell_fmt = (
                            CellFormat.from_props(user_entered_fmt)
                            if user_entered_fmt
                            else None
                        )
                        if cell_fmt:
                            # Horizontal alignment
                            if h_align := cell_fmt.horizontalAlignment:
                                attrs[CellAttribute.TEXT_ALIGN.value] = h_align.lower()

                            # Vertical alignment
                            if v_align := cell_fmt.verticalAlignment:
                                attrs[CellAttribute.VERTICAL_ALIGN.value] = (
                                    v_align.lower()
                                )

                            # Border
                            if border := get_border_data(cell_fmt.borders):
                                attrs[CellAttribute.BORDER.value] = border

                            text_fmt = cell_fmt.textFormat
                            if text_fmt:
                                # Font style
                                attrs[CellAttribute.TEXT_STYLE.value] = get_font_style(
                                    text_fmt, is_excel=False
                                )

                                # Font size
                                if text_fmt.fontSize:
                                    attrs[CellAttribute.FONT_SIZE.value] = (
                                        text_fmt.fontSize
                                    )
                                # Font name
                                if text_fmt.fontFamily in ALLOWED_FONTS:
                                    attrs[CellAttribute.FONT.value] = (
                                        text_fmt.fontFamily
                                    )

                                # Color
                                color = text_fmt.foregroundColor
                                if color and color != GSColor(0, 0, 0, 1):
                                    hex_color = color.toHex()
                                    attrs[CellAttribute.COLOR.value] = (
                                        hex_color
                                        if len(hex_color) == 7
                                        else hex_color[:-2]
                                    )

                                # Link
                                if text_fmt.link:
                                    attrs[CellAttribute.LINK.value] = text_fmt.link.uri
                            # Background color
                            if (
                                bgcolor := cell_fmt.backgroundColor
                            ) and bgcolor != GSColor(1, 1, 1, 1):
                                hex_color = bgcolor.toHex()
                                attrs[CellAttribute.BACKGROUND_COLOR.value] = (
                                    hex_color if len(hex_color) == 7 else hex_color[:-2]
                                )

                            # Wrap
                            if wrap := cell_fmt.wrapStrategy:
                                if wrap == "WRAP":
                                    attrs[CellAttribute.LINE_WRAP.value] = (
                                        LineWrap.WRAP.value
                                    )
                                elif wrap == "OVERFLOW_CELL":
                                    attrs[CellAttribute.LINE_WRAP.value] = (
                                        LineWrap.OVERFLOW.value
                                    )
                                elif wrap == "CLIP":
                                    attrs[CellAttribute.LINE_WRAP.value] = (
                                        LineWrap.TRUNCATE.value
                                    )

                        output_type = OutputType.EXECUTE_RESULT

                        if user_entered_value := cell_value.get("userEnteredValue"):
                            _user_entered_data_type, user_entered_code = next(
                                iter(user_entered_value.items())
                            )
                        else:
                            user_entered_code = None

                        traceback = None
                        ename = None
                        number_format = None

                        formatted_value = cell_value.get("formattedValue")

                        if effective_value_dict := cell_value.get("effectiveValue"):
                            if error := effective_value_dict.get("errorValue"):
                                traceback = error.get("message", "")
                                output_type = OutputType.ERROR
                                ename = formatted_value

                            effective_value = next(iter(effective_value_dict.values()))
                        else:
                            effective_value = None

                        if effective_format := cell_value.get("effectiveFormat"):
                            cell_num_format = effective_format.get("numberFormat", {})
                            if cell_num_format:
                                # TODO: add support for formatting patterns (cell_num_format['pattern'])
                                num_format_type = cell_num_format["type"]
                                if num_format_type == "PERCENT":
                                    number_format = NumberFormat.PERCENTAGE.value
                                elif num_format_type == "DATE":
                                    number_format = (
                                        f"{NumberFormat.DATE.value}-MM/dd/yyyy"
                                    )
                                elif num_format_type == "DATE_TIME":
                                    number_format = (
                                        f"{NumberFormat.DATE.value}-MM/dd/yyyy hh:mm:ss"
                                    )
                                elif num_format_type == "TIME":
                                    number_format = (
                                        f"{NumberFormat.DATE.value}-hh:mm:ss"
                                    )
                                elif num_format_type == "CURRENCY":
                                    number_format = NumberFormat.MONEY.value
                                elif num_format_type == "NUMBER":
                                    number_format = NumberFormat.FLOAT.value
                                else:
                                    number_format = NumberFormat.CUSTOM.value

                        if note := cell_value.get("note"):
                            attrs[CellAttribute.NOTE.value] = note

                        if number_format:
                            attrs[CellAttribute.NUMBER_FORMAT.value] = number_format

                        sheet_cell = SheetCell(
                            addr,
                            output=Output(
                                data={JSON_MIME_KEY: effective_value}
                                if output_type != OutputType.ERROR
                                else None,
                                execution_count=-1,
                                metadata=None,
                                output_type=output_type,
                                name=None,
                                text=None,
                                ename=ename,
                                evalue=None,
                                traceback=traceback,
                            ),
                            raw_code=str(user_entered_code),
                            attributes=attrs if attrs else None,
                        )
                        new_tyne_content.sheets.set(addr, sheet_cell)

            # Merged cells
            for merge in sh.get("merges", []):
                start_cell = cast(
                    SheetCell,
                    new_tyne_content.sheets.get(
                        Address(
                            merge["startColumnIndex"], merge["startRowIndex"], sheet.id
                        ),
                        True,
                    ),
                )
                if start_cell.attributes is None:
                    start_cell.attributes = {}
                cell_attrs = start_cell.attributes

                cell_attrs[CellAttribute.ROW_SPAN.value] = (
                    merge["endRowIndex"] - merge["startRowIndex"]
                )
                cell_attrs[CellAttribute.COL_SPAN.value] = (
                    merge["endColumnIndex"] - merge["startColumnIndex"]
                )

        return await self.new_tyne(  # type: ignore
            session, user, name=wb.title, content=new_tyne_content
        )

    async def import_xlsx(
        self, session: Session, file_buffer: bytes, file_name: str, user: User
    ) -> TyneModelWithContent:
        def replace_xlfn(matchobj: re.Match) -> str:
            assert matchobj.group(0)
            return matchobj.group(0)[6:].upper()

        bytes_io_buffer = BytesIO(file_buffer)
        try:
            wb_formulas = load_workbook(bytes_io_buffer)
            wb_values = load_workbook(bytes_io_buffer, data_only=True)
        except InvalidFileException:
            raise HTTPError(400, reason="Invalid XLSX file")
        except Exception:
            raise HTTPError(400, reason="Import error")

        theme_colors = get_theme_colors(wb_formulas)
        new_tyne_content = TyneContent.empty()
        for sheet_id, sheet_name in enumerate(wb_formulas.sheetnames):
            if sheet_id == 0:
                new_tyne_content.sheets.rename_sheet(0, sheet_name)
                sheet = new_tyne_content.sheets.sheets[sheet_id]
            else:
                name, sheet = new_tyne_content.sheets.new_sheet(sheet_name)

            sheet_values = wb_values[sheet_name]
            sheet.grid_size = calc_grid_size(
                sheet_values.max_column, sheet_values.max_row
            )

            sheet.attributes[SheetAttribute.ROWS_SIZES.value] = {
                str(i - 1): round(row_height.height)
                for i, row_height in sheet_values.row_dimensions.items()
                if row_height.customHeight
            }

            sheet.attributes[SheetAttribute.COLS_SIZES.value] = {
                str(col_width.min - 1): round(
                    col_width.width * OPENPYXL_COL_PIXELS_PER_UNIT
                )
                for col_width in sheet_values.column_dimensions.values()
                if col_width.customWidth
            }

            for row in zip(wb_formulas[sheet_name], sheet_values):
                for cell in zip(*row):
                    code = ""
                    number_format = None
                    ename = None
                    output_type = OutputType.EXECUTE_RESULT
                    cell_value = cell[0].value
                    attrs = {}

                    addr = Address(cell[0].column - 1, cell[0].row - 1, sheet.id)
                    if cell[0].data_type == TYPE_FORMULA:
                        if not isinstance(cell_value, str):
                            if not hasattr(cell_value, "text"):
                                continue
                            cell_value = cell_value.text
                        code = re.subn(XLFN_RE, replace_xlfn, cell_value)[0].replace(
                            "_xlpm.", ""
                        )
                        output_type = OutputType.EXECUTE_RESULT

                    if cell[1].data_type == TYPE_ERROR:
                        output_type = OutputType.ERROR
                        ename = cell[1].value
                        value = None
                    else:
                        value = cell[1].value

                    if isinstance(value, datetime):
                        number_format_lst = cell[1].number_format.split(" ")
                        number_format_lst[0] = number_format_lst[0].replace("m", "M")
                        if len(number_format_lst) == 1:
                            # Date
                            value = SpreadsheetDate(
                                value.replace(tzinfo=SpreadsheetTime.TZ_INFO).date()
                            )
                        else:
                            # Date/time
                            number_format_lst[1] = number_format_lst[1].replace(
                                "AM/PM", "a"
                            )
                            value = SpreadsheetDateTime(
                                value.replace(tzinfo=SpreadsheetTime.TZ_INFO)
                            )
                        number_format = f'date-{" ".join(number_format_lst)}'

                    elif isinstance(value, time):
                        value = SpreadsheetTime(
                            value.replace(tzinfo=SpreadsheetTime.TZ_INFO)
                        )
                        number_format = (
                            f"date-{cell[1].number_format.replace('AM/PM', 'a')}"
                        )
                    elif isinstance(value, str):
                        value = unescape(value)
                    elif cell[1].data_type == TYPE_NUMERIC:
                        if cell[1].number_format.endswith("%"):
                            number_format = NumberFormat.PERCENTAGE.value
                        elif cell[1].number_format.startswith('_("$'):
                            number_format = NumberFormat.MONEY.value

                    if number_format:
                        attrs[CellAttribute.NUMBER_FORMAT.value] = number_format

                    if note := cell[0].comment:
                        attrs[CellAttribute.NOTE.value] = note.text

                    # Border
                    if border := get_border_data(cell[1].border):
                        attrs[CellAttribute.BORDER.value] = border

                    if font := cell[1].font:
                        # Font style
                        if font_style := get_font_style(font, is_excel=True):
                            attrs[CellAttribute.TEXT_STYLE.value] = font_style
                        # Font size
                        if font.sz:
                            attrs[CellAttribute.FONT_SIZE.value] = font.sz
                        # Font name
                        if font.name in ALLOWED_FONTS:
                            attrs[CellAttribute.FONT.value] = font.name

                        # Color
                        if color := get_openpyxl_color(font.color, theme_colors):
                            attrs[CellAttribute.COLOR.value] = color

                    align = cell[1].alignment

                    # Horizontal alignment
                    if (h_align := align.horizontal) and h_align != "general":
                        attrs[CellAttribute.TEXT_ALIGN.value] = h_align

                    # Vertical alignment
                    if v_align := align.vertical:
                        if v_align == "top":
                            attrs[CellAttribute.VERTICAL_ALIGN.value] = (
                                VerticalAlign.TOP.value
                            )
                        elif v_align == "center":
                            attrs[CellAttribute.VERTICAL_ALIGN.value] = (
                                VerticalAlign.MIDDLE.value
                            )
                        elif v_align == "bottom":
                            attrs[CellAttribute.VERTICAL_ALIGN.value] = (
                                VerticalAlign.BOTTOM.value
                            )

                    # Line wrap
                    line_wrap = align.wrapText
                    if line_wrap:
                        attrs[CellAttribute.LINE_WRAP.value] = LineWrap.WRAP.value

                    # Background color
                    if bgcolor := get_openpyxl_color(
                        cell[1].fill.fgColor, theme_colors
                    ):
                        attrs[CellAttribute.BACKGROUND_COLOR.value] = bgcolor

                    # Link
                    link = cell[1].hyperlink
                    if link:
                        attrs[CellAttribute.LINK.value] = link.target

                    sheet_cell = SheetCell(
                        addr,
                        output=Output(
                            data={JSON_MIME_KEY: value}  # type: ignore
                            if output_type != OutputType.ERROR
                            else None,
                            execution_count=-1,
                            metadata=None,
                            output_type=output_type,
                            name=None,
                            text=None,
                            ename=ename,
                            evalue="",
                            traceback=None,
                        ),
                        raw_code=str(code) if code else str(value) if value else "",
                        attributes=attrs if attrs else None,
                    )
                    new_tyne_content.sheets.set(addr, sheet_cell)
            for merged_range in sheet_values.merged_cells:
                start_coord = merged_range.left[0]
                start_cell = cast(
                    SheetCell,
                    new_tyne_content.sheets.get(
                        Address(start_coord[1] - 1, start_coord[0] - 1, sheet.id), True
                    ),
                )
                if start_cell.attributes is None:
                    start_cell.attributes = {}
                cell_attrs = start_cell.attributes

                cell_attrs[CellAttribute.ROW_SPAN.value] = merged_range.size["rows"]
                cell_attrs[CellAttribute.COL_SPAN.value] = merged_range.size["columns"]

        return await self.new_tyne(session, user, file_name, content=new_tyne_content)  # type: ignore
