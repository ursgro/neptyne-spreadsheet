import csv
from io import BytesIO, StringIO
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import (
    TYPE_BOOL,
    TYPE_ERROR,
    TYPE_FORMULA,
    TYPE_NUMERIC,
    TYPE_STRING,
)
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.styles import Border as ExcelBorder
from openpyxl.styles.colors import Color
from openpyxl.styles.numbers import (
    FORMAT_DATE_DATETIME,
    FORMAT_DATE_TIME1,
    FORMAT_DATE_YYYYMMDD2,
    FORMAT_PERCENTAGE,
)
from tornado.web import HTTPError

from neptyne_kernel.cell_address import Address
from neptyne_kernel.expression_compiler import format_col
from neptyne_kernel.mime_handling import output_to_value
from neptyne_kernel.mime_types import JSON_MIME_KEY
from neptyne_kernel.neptyne_protocol import (
    CellAttribute,
    LineWrap,
    SheetAttribute,
    TextStyle,
)
from neptyne_kernel.spreadsheet_datetime import (
    excel2date,
    excel2datetime,
    excel2time,
)
from neptyne_kernel.spreadsheet_error import SPREADSHEET_ERRORS_STR
from neptyne_kernel.tyne_model.jupyter_notebook import Output
from neptyne_kernel.tyne_model.sheet import TyneSheets
from server.tyne_import import OPENPYXL_COL_PIXELS_PER_UNIT


class TyneExportMixin:
    def export_xlsx(self, tyne_sheets: TyneSheets, properties: dict[str, Any]) -> bytes:
        wb = Workbook()
        if properties and "sheetsOrder" in properties:
            sheet_keys = properties["sheetsOrder"]
        else:
            sheet_keys = tyne_sheets.sheets.keys()
        first_sheet = True
        for sheet_key in sheet_keys:
            try:
                sheet = tyne_sheets.sheets[sheet_key]
            except KeyError:
                # sheetsOrder may contain a deleted sheet ID
                continue
            ranges_to_merge = []
            if first_sheet:
                ws = wb.active
                ws.title = sheet.name
                first_sheet = False
            else:
                ws = wb.create_sheet(title=sheet.name)
            for addr, cell in sheet.cells.items():
                wc = ws.cell(column=addr.column + 1, row=addr.row + 1)
                if attrs := cell.attributes:
                    # Horizontal alignment
                    h_align = attrs.get(CellAttribute.TEXT_ALIGN.value)
                    # Vertical alignment
                    v_align = attrs.get(CellAttribute.VERTICAL_ALIGN.value)
                    if v_align == "middle":
                        v_align = "center"
                    # Wrap
                    is_wrapped = (
                        attrs.get(CellAttribute.LINE_WRAP.value) == LineWrap.WRAP.value
                    )

                    wc.alignment = Alignment(
                        horizontal=h_align, vertical=v_align, wrapText=is_wrapped
                    )

                    # Border
                    if borders := set(attrs.get("border", "").split()):
                        wc.border = ExcelBorder(
                            **{
                                side: Side(style="thin")
                                for side in ("top", "bottom", "left", "right")
                                if f"border-{side}" in borders
                            }
                        )

                    # Font style
                    font_styles = (
                        set(font_style.split())
                        if (font_style := attrs.get(CellAttribute.TEXT_STYLE.value))
                        else set()
                    )
                    # Color
                    font_color = (
                        Color(rgb="00" + color[1:])
                        if (color := attrs.get(CellAttribute.COLOR.value))
                        else None
                    )

                    # Font size
                    font_size = attrs.get(CellAttribute.FONT_SIZE.value)
                    # Font family
                    font_name = attrs.get(CellAttribute.FONT.value)
                    wc.font = Font(
                        sz=font_size,
                        name=font_name,
                        b=TextStyle.BOLD.value in font_styles,
                        i=TextStyle.ITALIC.value in font_styles,
                        u="single"
                        if (TextStyle.UNDERLINE.value in font_styles)
                        else None,
                        color=font_color,
                    )

                    # Link
                    if link := attrs.get(CellAttribute.LINK.value):
                        wc.hyperlink = link

                    # Background color
                    if bg_color := attrs.get(CellAttribute.BACKGROUND_COLOR.value):
                        wc.fill = PatternFill(
                            patternType="solid", fgColor=Color(rgb=bg_color[1:])
                        )

                    # Merged cells
                    row_span = attrs.get(CellAttribute.ROW_SPAN.value, 0)
                    col_span = attrs.get(CellAttribute.COL_SPAN.value, 0)
                    if row_span or col_span:
                        start_addr = cell.cell_id
                        ranges_to_merge.append(
                            dict(
                                start_row=start_addr.row + 1,
                                start_column=start_addr.column + 1,
                                end_row=start_addr.row + row_span,
                                end_column=start_addr.column + col_span,
                            )
                        )
                    # Note
                    if note := attrs.get(CellAttribute.NOTE.value):
                        wc.comment = Comment(text=note, author=None)

                if cell.raw_code.startswith("="):
                    wc.data_type = TYPE_FORMULA
                    wc.value = cell.raw_code

                number_format: str = ""
                value: Any = None
                data_type: str | None = None

                data = (
                    cell.output.data.get(JSON_MIME_KEY)
                    if isinstance(cell.output, Output)
                    and cell.output
                    and cell.output.data
                    else cell.output
                )

                if data is None:
                    data = cell.raw_code
                if isinstance(data, bool):
                    value = data
                    data_type = TYPE_BOOL
                elif isinstance(data, int | float):
                    value = data
                    fmt = cell.attributes.get("numberFormat") if cell.attributes else ""
                    if fmt:
                        if fmt.startswith("date"):
                            data_fmt = None
                            if len(fmt) > 5:
                                if fmt[5] == "-":
                                    data_fmt = fmt[6:]
                            if abs(data) < 1:
                                value = excel2time(data)
                                number_format = (
                                    data_fmt if data_fmt else FORMAT_DATE_TIME1
                                )
                            elif isinstance(data, int):
                                value = excel2date(data)
                                number_format = (
                                    data_fmt if data_fmt else FORMAT_DATE_YYYYMMDD2
                                )
                            else:
                                value = excel2datetime(data)
                                number_format = (
                                    data_fmt if data_fmt else FORMAT_DATE_DATETIME
                                )

                        elif fmt.startswith("percentage"):
                            number_format = FORMAT_PERCENTAGE
                    data_type = TYPE_NUMERIC
                elif isinstance(data, str):
                    if data in SPREADSHEET_ERRORS_STR:
                        value = data
                        data_type = TYPE_ERROR
                    else:
                        value = data
                        data_type = TYPE_STRING

                if number_format:
                    wc.number_format = number_format
                if wc.data_type != TYPE_FORMULA:
                    wc.value = value
                    wc.data_type = data_type

            for cells_to_merge in ranges_to_merge:
                ws.merge_cells(**cells_to_merge)

            sheet_attrs = sheet.attributes

            # Row heights
            if row_sizes := sheet_attrs.get(SheetAttribute.ROWS_SIZES.value):
                for i, height in row_sizes.items():
                    ws.row_dimensions[int(i) + 1].height = height

            # Column widths
            if col_sizes := sheet_attrs.get(SheetAttribute.COLS_SIZES.value):
                for j, width in col_sizes.items():
                    ws.column_dimensions[format_col(int(j))].width = (
                        width / OPENPYXL_COL_PIXELS_PER_UNIT
                    )

        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)
        return virtual_workbook.getvalue()

    def export_csv(self, tyne_sheets: TyneSheets, sheet_id: int) -> str:
        ft = StringIO()
        w = csv.writer(ft, lineterminator="\n")
        sheet = tyne_sheets.sheets.get(sheet_id)
        if not sheet:
            raise HTTPError(400, reason=f"Sheet {sheet_id} doesn't exist")
        for i in range(sheet.grid_size[1]):
            row: list[Any] = []
            for j in range(sheet.grid_size[0]):
                cell = sheet.cells.get(Address(j, i, sheet_id))
                if cell and cell.output:
                    if hasattr(cell.output, "ename") and cell.output.ename:
                        row.append(cell.output.ename)
                    else:
                        value = (
                            output_to_value(cell.output.data)
                            if isinstance(cell.output, Output)
                            else cell.output
                        )
                        row.append(str(value) if value is not None else None)
                else:
                    row.append(None)
            w.writerow(row)
        return ft.getvalue()
