import asyncio
import contextlib
import json
import os
import random
import re
import string
from concurrent.futures import Executor, ProcessPoolExecutor
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, ContextManager, Generator

import aiohttp
import gspread
import gspread_asyncio
import openpyxl
import tiktoken
from gcloud.aio.storage import Storage
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openai import AsyncClient, RateLimitError
from openai.types.chat import (
    ChatCompletionSystemMessageParam,
    ChatCompletionUserMessageParam,
)
from openpyxl.chartsheet import Chartsheet
from tiktoken import Encoding
from tornado.web import HTTPError, RequestHandler

from neptyne_kernel.cell_address import Address

DO_UPLOADS = True

COLUMN_SEPARATOR = "|"
CELL_ASSIGNMENT = "←"

CELL_LIMIT = 6000
TOKEN_LIMIT = 1024 * 64

MODEL = "gpt-4o"

LINTER_SERVER_HOST = os.getenv("EXCELINT_SERVICE_SERVICE_HOST", "localhost")
LINTER_SERVER_PORT = os.getenv("EXCELINT_SERVICE_SERVICE_PORT", "3003")
LINTER_SERVER = f"http://{LINTER_SERVER_HOST}:{LINTER_SERVER_PORT}/upload"

LEVELS = ["high", "medium", "low"]

PROBLEMS = [
    (
        "Relative and Absolute References in Formulas",
        "A formula uses relative references when it should use absolute references or vice versa",
    ),
    (
        "Mixed Data Types",
        "Cells with the same functions contain different data types, "
        "numbers vs strings or numbers with different units or suspiciously different precisions",
    ),
    (
        "Reference to the wrong column or row",
        "Cell refers to a column or row different from its own, while similar cells refer to the correct column or row",
    ),
    (
        "Hardcoded Numbers in Formulas",
        "Either because a calculation can be simplified by resolving the formula or if "
        "a hardcoded number appears elsewhere in the spreadsheet",
    ),
    (
        "Incorrect Footnote References",
        "Footnotes or comments that are meant to provide additional information but are not correctly "
        "referenced within the sheet, leading to confusion or misinformation.",
    ),
    (
        "Overcomplicated Formulas",
        "Formulas whose complexity makes them hard to read and that can be simplified or broken down into smaller "
        "parts by using named ranges or helper cells.",
    ),
    (
        "Poorly Organized Data",
        "Data that is organized in a way that makes it hard to understand, for example no clear headers, "
        "separation of tables or switching between orientation.",
    ),
]

SAMPLE_OUTPUT = {
    # see: https://docs.google.com/spreadsheets/d/1mPbLooOQ5ur4-FMqNbOD16F3OoKaiWU5cs5epGn2CUA/edit#gid=141914970
    "description": "Description in some detail about what is going on in this spreadsheet.",
    "summary": "A ~50 word summary of the description.",
    "subtables": {"Sheet1": ["A2:B5", "D2:E5"], "Sheet2": ["A1:B3"]},
    "calculations": [
        "The top half of sheet1 calculates the return on investment "
        "for a given set of investments",
        "The bottom half of sheet1 lists three different scenarios ",
        "sheet2 contains the assumptions",
    ],
    "problems": [
        {
            "sheet": "Sheet1",
            "address": "C1",
            "example": "C1 (fraction for Val1) contains the formula =B1/B4, "
            "but B4 represents the total value, so it should be an absolute reference",
            "problem": "Relative and Absolute References in Formulas",
            "fix": "Change the formula to =B1/B$4",
            "severity": "medium",
        },
        {
            "sheet": "Sales",
            "address": "D4",
            "example": "D4 (units sold) contains 25.3212, but the rest of the D column contains "
            "rounded numbers",
            "problem": "Mixed Data Types",
            "fix": "Check the value of D4 and possibly change to 25",
            "severity": "medium",
        },
        {
            "sheet": "Products",
            "address": "E5",
            "example": "E5 (total) contains =C5*D6, but E6 contains =C6*D6, suggesting that E5 refers to the wrong cell",
            "problem": "Reference to the wrong column or row",
            "fix": "Change the formula in E5 to =C5*D5",
            "severity": "high",
        },
        {
            "sheet": "Overview",
            "address": "C5",
            "example": "C5 (total price) contains =SUM(C2:C4) * 2.31, but it is unclear where the 2.31 comes from",
            "problem": "Hardcoded Numbers in Formulas",
            "fix": "Move the 2.31 to a separate cell and refer to it in the formula",
            "severity": "medium",
        },
    ],
}

PROMPT_PATTERN = (
    "Find problems and suggest solutions for the following spreadsheet.\n"
    f"The format is name:<sheetname> followed by the cells, separated by {COLUMN_SEPARATOR}. "
    f"Each cell contains the cell address followed by a {CELL_ASSIGNMENT} followed by the contents\n"
    "Here is the spreadsheet:\n"
    "{spreadsheet}"
    "{static}"
    "\n\n"
    "You are looking for problems like this:\n"
    "{problems}\n\n"
    "Given this spreadsheet, start by finding subtables. "
    "Scan from top to bottom, left to write. Create a dict with as key the name of the sheet and as "
    "values a list of ranges for the tables in excel notation.\n"
    "Then create an overview of what the spreadsheet is doing. For each sheet or large sub-table, "
    "describe what is being calculated and how.\n"
    "Then look find problems and suggested fixes. For each problem identify where this happens "
    "(sheet, cell address), what the problem is, what a fix could look like and a concrete example of the "
    'problem, something like "in cell A1 the formula is SUM(A2:A5) but in C1 it says SUM(C2:C6)". '
    "In the example always make sure you mention the actual value of the cell that causes the problem. "
    "Mark each problem's severity: high - the wrong values are being calculated. medium - works as is, "
    "but small changes could lead to problems. low - not an immediate concern, but cleaning this up "
    "could help readability and understanding.\n"
    "Finally put it all together in one json document with keys {json_keys}.\n"
    "It should look like this:\n"
    "{example}"
)

HTML = """<html>
<head>
<style>
  body {{ font-family: Arial, sans-serif; margin: 20px; }}
  h2, h3 {{ color: #333; }}
  .description, .calculation, .problem {{ margin-bottom: 20px; }}
  .problem-details {{ margin-left: 20px; }}
  .severity-high {{ color: red; }}
  .severity-medium {{ color: orange; }}
  img {{ display: block; margin-top: 10px; }}
  ul {{ list-style-type: none; padding: 0; }}
  li {{ margin-bottom: 10px; }}
  li:before {{ content: '•'; color: #333; display: inline-block; width: 1em; margin-left: -1em; }}
</style>
<title>{title}</title>
</head>
<body>
<h2>{title}</h2>
<div class="description"><strong>Description:</strong> {description}</div>
<h3>Calculations:</h3>
<ul>
{calculations}
</ul>
<h3>Problems:</h3>
{problems}
</body></html>
"""

CALCULATION_HTML = "<li class='calculation'>{calculation}</li>"

PROBLEM_HTML = """<div class="problem {severity_class}">
<strong>{sheet} {address}:</strong> {problem}
<div class="problem-details">
<strong>Fix:</strong> {fix}
</div>
</div>"""


class FileTooLargeError(ValueError):
    def __str__(self) -> str:
        return (
            "Your document is too big for this preview version of the sheet bot. "
            "Please try something smaller"
        )


async def upload_to_gcs(data: bytes | str, filename: str) -> None:
    if not DO_UPLOADS:
        return
    async with aiohttp.ClientSession() as http_session:
        storage = Storage(session=http_session)  # type: ignore
        await storage.upload(
            "sheet-bot-uploads",
            filename,
            data,
        )


def format_prompt(
    spreadsheet: str,
    static_analysis: dict[str, Any],
    prompt_pattern: str = PROMPT_PATTERN,
) -> str:
    return prompt_pattern.format(
        spreadsheet=spreadsheet,
        problems="\n".join(f"- {' - '.join(p)}" for p in PROBLEMS),
        static="\n\nStatic analysis have found the following problems:\n"
        + json.dumps(static_analysis, indent=4)
        if static_analysis
        else "",
        example=json.dumps(SAMPLE_OUTPUT, indent=4),
        json_keys=", ".join(SAMPLE_OUTPUT.keys()),
    )


TIKTOKEN_ENCODING = None


def get_tiktoken_model(model: str = MODEL) -> Encoding:
    global TIKTOKEN_ENCODING
    if TIKTOKEN_ENCODING is None:
        TIKTOKEN_ENCODING = tiktoken.encoding_for_model(model)
    return TIKTOKEN_ENCODING


async def run_prompt(
    spreadsheet: str,
    static_analysis: dict[str, Any],
    prompt_pattern: str = PROMPT_PATTERN,
) -> dict[str, Any]:
    client = AsyncClient()
    prompt = format_prompt(spreadsheet, static_analysis, prompt_pattern)
    token_count = len(get_tiktoken_model().encode(prompt))
    if token_count > TOKEN_LIMIT:
        raise FileTooLargeError(f"token count: {token_count}")
    system_message: ChatCompletionSystemMessageParam = {
        "role": "system",
        "content": "You are a efficient assistant helping a user with their spreadsheet. "
        "You find problems and suggest solutions",
    }
    user_message: ChatCompletionUserMessageParam = {
        "role": "user",
        "content": prompt,
    }
    try:
        completion = await client.chat.completions.create(
            model=MODEL,
            messages=[system_message, user_message],
            response_format={"type": "json_object"},
            temperature=0.0,
        )
    except RateLimitError as e:
        raise HTTPError(413, f"OpenAI API rate limit exceeded: {e}")

    reply = completion.choices[0].message.content
    if reply is None:
        raise ValueError("Prompt failed to generate a response")
    res = json.loads(reply)
    res["prompt"] = prompt
    if completion.usage:
        res["promptTokens"] = completion.usage.prompt_tokens
        res["completionTokens"] = completion.usage.completion_tokens
        res["totalTokens"] = completion.usage.total_tokens

    def priority(problem: dict[str, Any]) -> int:
        try:
            return LEVELS.index(problem["severity"])
        except (IndexError, KeyError, ValueError):
            return len(LEVELS)

    if "problems" in res:
        res["problems"] = sorted(res["problems"], key=priority)
    return res


def format_cell(address: Address, value: str | None) -> str:
    if cell_value_is_empty(value):
        return ""
    cell_value = address.to_a1() + "←" + str(value).replace("←", "<-")
    cell_value = cell_value.replace("\\", "\\\\")
    cell_value = cell_value.replace("\n", "\\n")
    cell_value = cell_value.replace(COLUMN_SEPARATOR, "¦")
    return cell_value


class InvalidExcelFile(Exception):
    pass


Grids = list[tuple[str, list[list[str | None]]]]


async def excel_to_grids(
    blob_or_filename: BytesIO | str, executor: Executor | None = None
) -> Grids:
    if executor is None:
        with ProcessPoolExecutor() as executor:
            return await excel_to_grids(blob_or_filename, executor=executor)
    return await asyncio.get_running_loop().run_in_executor(
        executor, exec_excel_to_grids, blob_or_filename
    )


def cell_value_is_empty(value: Any) -> bool:
    return value is None or value == ""


def exec_excel_to_grids(
    blob_or_filename: BytesIO | str, cell_limit: int = CELL_LIMIT
) -> Grids:
    try:
        workbook = openpyxl.load_workbook(blob_or_filename)
    except Exception as e:
        raise InvalidExcelFile(f"Failed to load the workbook: {e}")
    sheets = []
    total_cells = 0
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        if isinstance(worksheet, Chartsheet):
            continue
        rows = []
        for row in range(1, worksheet.max_row + 1):
            row_values = []

            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                if not cell_value_is_empty(cell.value):
                    total_cells += 1
                    if cell_limit and total_cells > cell_limit:
                        raise FileTooLargeError(f"cell count: {total_cells}")
                row_values.append(cell.value)
            rows.append(row_values)
        sheets.append((sheet, rows))
    return sheets


async def sheets_to_grids(
    url: str, credentials: Credentials, executor: Executor
) -> Grids:
    if m := re.search(r"file/d/([^/]+)", url):
        # hosted excel file:
        file_id = m.group(1)
        drive_service = build("drive", "v3", credentials=credentials)
        request = drive_service.files().get_media(fileId=file_id)
        blob = BytesIO()
        downloader = MediaIoBaseDownload(blob, request)
        done = False
        while done is False:
            _, done = downloader.next_chunk()
        blob.seek(0)
        return await excel_to_grids(blob, executor)
    try:
        gc = await gspread_asyncio.AsyncioGspreadClientManager(
            lambda: credentials
        ).authorize()
        wb = await gc.open_by_url(url)
    except gspread.exceptions.NoValidUrlKeyFound:
        raise ValueError("We could not read the Google Sheets document")
    except gspread.exceptions.APIError:
        raise ValueError("We could not read that file as a valid Excel document")
    wb_data = await wb.fetch_sheet_metadata(
        {
            "fields": "sheets",
            "includeGridData": True,
        }
    )

    sheets = []
    for sh in wb_data["sheets"]:
        rows = []
        data = sh["data"][0]
        for row in data.get("rowData", []):
            if row and (row_values := row.get("values")):
                cells = []
                for cell in row_values:
                    if not cell:
                        cell_value = None
                    else:
                        if isinstance(cell, str):
                            continue
                        if user_entered_value := cell.get("userEnteredValue"):
                            _, cell_value = next(iter(user_entered_value.items()))
                        else:
                            cell_value = None
                    cells.append(cell_value)
                rows.append(cells)
            else:
                rows.append([])
        sheets.append((sh["properties"]["title"], rows))
    return sheets


def grids_to_text(grids: Grids) -> str:
    sheets = []
    for grid_idx, (sheet_name, grid) in enumerate(grids):
        text_representation = ""
        for row_idx, row in enumerate(grid):
            row_values = [
                format_cell(Address(col_idx, row_idx, grid_idx), cell)
                for col_idx, cell in enumerate(row)
            ]
            while row_values and row_values[-1] == "":
                row_values.pop()
            text_representation += COLUMN_SEPARATOR.join(row_values) + "\n"
        text_representation = text_representation.rstrip("\n")
        if text_representation:
            sheets.append("name: " + sheet_name + "\n" + text_representation)
    return "\n\n".join(sheets).strip()


async def call_linter(
    file_body: bytes, file_name: str, content_type: str
) -> dict[str, Any] | None:
    data = aiohttp.FormData()
    data.add_field(
        "xlsfile",
        file_body,
        filename=file_name,
        content_type=content_type,
    )

    async with aiohttp.ClientSession() as session:
        async with session.post(LINTER_SERVER, data=data) as response:
            if response.status == 200:
                linted = await response.json()

                def sheet_summary(sheet: dict) -> list[dict[str, Any]]:
                    def fix_summary(fix: dict) -> dict[str, Any] | None:
                        t = fix["_analysis"].get("t")
                        if not t:
                            return None
                        return {
                            "score": fix["_score"],
                            "references": [a["print_formula"] for a in t["analysis"]],
                            "issues": t["classification"],
                        }

                    return [
                        summary
                        for fix in sheet["analysis"]["proposed_fixes"]
                        if (summary := fix_summary(fix))
                    ]

                summary = {
                    sheet_name: summary
                    for sheet_name, sheet in linted[0]["sheets"].items()
                    if (summary := sheet_summary(sheet))
                }
                return summary
            else:
                return None


@contextlib.contextmanager
def timer(name: str) -> Generator:
    start = datetime.utcnow()
    print("Starting", name)
    yield
    print(f"{name} took {datetime.utcnow() - start}")


class SheetLinterHandler(RequestHandler):
    file_name: str
    executor: Executor

    def initialize(self, executor: Executor) -> None:
        self.executor = executor

    def timer(self, name: str) -> ContextManager[None]:
        return timer(f"{self.file_name}:{name}")

    async def gather_timed(self, *coroutines: Any) -> tuple:
        async def timed(coroutine: Any) -> Any:
            with self.timer(coroutine.__name__):
                return await coroutine

        timed_coros = (timed(coroutine) for coroutine in coroutines)
        future = asyncio.gather(*timed_coros)
        try:
            return await future
        except Exception:
            future.cancel()
            raise

    async def get(self) -> None:
        # set up the form:
        self.write(
            """
            <html>
            <body>
            <form action="/api/sheet_linter" method="post" enctype="multipart/form-data">
            <input type="file" name="file1">
            <input type="submit" value="Upload">
            </form>
            </body>
            </html>
            """
        )

    async def handle_post(self) -> None:
        self.file_name = file_name = datetime.utcnow().strftime("%Y-%m-%d-") + "".join(
            random.choices(string.ascii_lowercase, k=5)
        )
        if sh := os.getenv("GIT_SHA"):
            file_name = sh + "/" + file_name
        problems: dict[str, Any] | None
        if authPayload := self.get_argument("authPayload", None):
            url = self.get_argument("url")
            oauth_access = json.loads(authPayload)
            expiry = (
                datetime.utcnow() + timedelta(seconds=oauth_access["expires_in"])
            ).isoformat()
            credentials = Credentials.from_authorized_user_info(
                {
                    **oauth_access,
                    "token": oauth_access["access_token"],
                    "expiry": expiry,
                    "refresh_token": None,
                    "client_id": None,
                    "client_secret": None,
                },
                scopes=[
                    "https://www.googleapis.com/auth/drive.file",
                ],
            )
            problems = {}
            with self.timer("sheets_to_grids"):
                grids = await sheets_to_grids(url, credentials, self.executor)
            with self.timer("upload_to_gcs"):
                await upload_to_gcs(grids_to_text(grids), file_name + ".sheets.txt")
        else:
            uploaded_file = self.request.files["file1"][0]

            try:
                _, grids, problems = await self.gather_timed(
                    upload_to_gcs(
                        uploaded_file["body"],
                        file_name + "-" + uploaded_file["filename"],
                    ),
                    excel_to_grids(BytesIO(uploaded_file["body"]), self.executor),
                    call_linter(
                        uploaded_file["body"],
                        uploaded_file["filename"],
                        uploaded_file["content_type"],
                    ),
                )
            except InvalidExcelFile as e:
                raise HTTPError(415, f"Invalid file: {e}")
            if problems is None:
                raise HTTPError(500, "Linter server failed to process file")

        with self.timer("run_prompt"):
            output = await run_prompt(grids_to_text(grids), problems)
        problems_formatted = [
            PROBLEM_HTML.format(
                sheet=problem["sheet"],
                address=problem["address"],
                problem=problem["problem"],
                fix=problem["fix"],
                severity_class=f"severity-{problem['severity']}",
            )
            for problem in output["problems"]
        ]

        calculations = "".join(
            [
                CALCULATION_HTML.format(calculation=calculation)
                for calculation in output["calculations"]
            ]
        )

        out_format = self.get_argument("format", "html")
        if out_format == "json":
            self.set_header("Content-Type", "application/json")
            self.write(output)
            await upload_to_gcs(json.dumps(output, indent=4), file_name + ".json")
        else:
            assert out_format == "html"
            self.set_header("Content-Type", "text/html")
            html = HTML.format(
                title="Analysis of: " + uploaded_file["filename"],
                description=output["description"],
                calculations=calculations,
                problems="".join(problems_formatted),
            )

            self.write(html)

    async def post(self) -> None:
        try:
            await self.handle_post()
            print("successfully linted:", self.file_name)
        except ValueError as e:
            if isinstance(e, FileTooLargeError):
                print("file too large:", self.file_name, e.args)
            out_format = self.get_argument("format", "html")
            if out_format == "json":
                self.set_header("Content-Type", "application/json")
                self.write({"error": str(e)})
            else:
                raise
